#!/usr/bin/env python3
"""
Excel generator for payroll system
Creates Excel files with Thai support using openpyxl
"""

import sys
import json
import base64
from io import BytesIO
from pathlib import Path

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print(json.dumps({"error": "openpyxl not installed. Run: pip install openpyxl"}))
    sys.exit(1)


def create_summary_excel(summaries: list) -> bytes:
    """Create summary Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "สรุปยอดจ่ายเงิน"
    
    # Styles
    header_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
    header_font = Font(bold=True, size=12)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Headers
    headers = ["ชื่อพนักงาน", "วันที่ทำงาน (วัน)", "ชั่วโมงทำงาน (ชม.)", "ค่าจ้างปกติ (บาท)", "หักมาสาย (บาท)", "รับเงินสุทธิ (บาท)"]
    header_row = ws.append(headers)
    
    # Style header row
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin
    
    # Add data
    for summary in summaries:
        ws.append([
            summary["name"],
            summary["totalDays"],
            summary["totalHours"],
            summary["basePay"],
            summary["totalPenalty"],
            summary["netPay"],
        ])
    
    # Style data rows
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=6):
        for cell in row:
            cell.border = thin
            cell.alignment = center
    
    # Auto-fit columns
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column].width = adjusted_width
    
    # Save to bytes
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_individual_excel(emp_name: str, records: list, period_text: str = "") -> bytes:
    """Create individual employee Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "เงินเดือน"
    
    # Styles
    header_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    # Headers
    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value=f"เงินเดือนพนักงานติดมันส์สาขาหาดใหญ่ {emp_name}").font = title_font
    
    ws.merge_cells("A2:G2")
    ws.cell(row=2, column=1, value="โทร: .......................................")
    
    ws.merge_cells("A3:C3")
    ws.cell(row=3, column=1, value=f"วันที่ {period_text}" if period_text else "วันที่ ........................")
    ws.cell(row=3, column=5, value="เงินสด")
    
    # Column headers
    col_headers = ["รหัสพนักงาน", "ชื่อ", "วันที่", "เวลาเข้างาน", "คิดเป็นเงิน", "ทิป", "รวมเงิน", "เบิกเงิน", "เหลือสุทธิ", "หมายเหตุ(เบิก)"]
    for col_num, header in enumerate(col_headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    # Data rows
    last_row = 5
    for row_idx, rec in enumerate(records, 6):
        last_row = row_idx
        date_val = rec.get("date", "")
        if hasattr(date_val, "strftime"):
            date_val = date_val.strftime("%d/%m/%Y")

        values = [
            rec.get("staffId", ""),
            rec.get("name", ""),
            date_val,
            rec.get("checkInTime", ""),
            rec.get("basePay", 0),
            rec.get("tip", 0),
            f"=E{row_idx}+F{row_idx}",
            rec.get("advance", ""),
            f"=IFERROR(G{row_idx}-H{row_idx},0)",
            rec.get("advanceNote", ""),
        ]

        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.border = thin
            if col_num >= 3:
                cell.alignment = center

    # Summary row
    summary_row = last_row + 1
    if last_row >= 6:
        ws.cell(row=summary_row, column=2, value=f"รวม {period_text}".strip()).alignment = center
        ws.cell(row=summary_row, column=5, value=f"=SUM(E6:E{last_row})").border = thin
        ws.cell(row=summary_row, column=6, value=f"=SUM(F6:F{last_row})").border = thin
        ws.cell(row=summary_row, column=7, value=f"=SUM(G6:G{last_row})").border = thin
        ws.cell(row=summary_row, column=8, value=f"=SUM(H6:H{last_row})").border = thin
        ws.cell(row=summary_row, column=9, value=f"=SUM(I6:I{last_row})").border = thin

    for c in range(1, 11):
        ws.cell(row=summary_row, column=c).border = thin

    # Empty rows
    for r in [summary_row + 1, summary_row + 2]:
        for c in range(1, 11):
            ws.cell(row=r, column=c).border = thin

    # Signature row
    sig_row = summary_row + 3
    ws.cell(row=sig_row, column=2, value="ลายเซ็นต์รับเงิน").alignment = center
    ws.cell(row=sig_row, column=8, value="วันที่").alignment = center
    for c in range(1, 11):
        ws.cell(row=sig_row, column=c).border = thin

    # Column widths
    col_widths = [14, 20, 14, 14, 12, 8, 12, 10, 12, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def create_all_employees_excel(employees_data: list, period_text: str = "") -> bytes:
    """Create Excel with all employees (each in separate sheet)"""
    wb = Workbook()
    
    # Styles
    header_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
    header_font = Font(bold=True, size=11)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )
    
    for idx, emp_data in enumerate(employees_data):
        emp_name = emp_data["emp_name"]
        records = emp_data["records"]
        sheet_title = str(emp_name).replace("/", "-").replace("\\", "-")[:31] or f"Employee_{idx+1}"
        
        if idx == 0:
            ws = wb.active
            ws.title = sheet_title
        else:
            ws = wb.create_sheet(title=sheet_title)
        
        # Headers
        ws.merge_cells("A1:G1")
        ws.cell(row=1, column=1, value=f"เงินเดือนพนักงานติดมันส์สาขาหาดใหญ่ {emp_name}").font = title_font
        
        ws.merge_cells("A2:G2")
        ws.cell(row=2, column=1, value="โทร: .......................................")
        
        ws.merge_cells("A3:C3")
        ws.cell(row=3, column=1, value=f"วันที่ {period_text}" if period_text else "วันที่ ........................")
        ws.cell(row=3, column=5, value="เงินสด")
        
        # Column headers
        col_headers = ["รหัสพนักงาน", "ชื่อ", "วันที่", "เวลาเข้างาน", "คิดเป็นเงิน", "ทิป", "รวมเงิน", "เบิกเงิน", "เหลือสุทธิ", "หมายเหตุ(เบิก)"]
        for col_num, header in enumerate(col_headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin

        # Data rows
        last_row = 5
        for row_idx, rec in enumerate(records, 6):
            last_row = row_idx
            date_val = rec.get("date", "")
            if hasattr(date_val, "strftime"):
                date_val = date_val.strftime("%d/%m/%Y")

            values = [
                rec.get("staffId", ""),
                rec.get("name", ""),
                date_val,
                rec.get("checkInTime", ""),
                rec.get("basePay", 0),
                rec.get("tip", 0),
                f"=E{row_idx}+F{row_idx}",
                rec.get("advance", ""),
                f"=IFERROR(G{row_idx}-H{row_idx},0)",
                rec.get("advanceNote", ""),
            ]

            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_num, value=val)
                cell.border = thin
                if col_num >= 3:
                    cell.alignment = center

        # Summary row
        summary_row = last_row + 1
        if last_row >= 6:
            ws.cell(row=summary_row, column=2, value=f"รวม {period_text}".strip()).alignment = center
            ws.cell(row=summary_row, column=5, value=f"=SUM(E6:E{last_row})").border = thin
            ws.cell(row=summary_row, column=6, value=f"=SUM(F6:F{last_row})").border = thin
            ws.cell(row=summary_row, column=7, value=f"=SUM(G6:G{last_row})").border = thin
            ws.cell(row=summary_row, column=8, value=f"=SUM(H6:H{last_row})").border = thin
            ws.cell(row=summary_row, column=9, value=f"=SUM(I6:I{last_row})").border = thin

        for c in range(1, 11):
            ws.cell(row=summary_row, column=c).border = thin

        for r in [summary_row + 1, summary_row + 2]:
            for c in range(1, 11):
                ws.cell(row=r, column=c).border = thin

        sig_row = summary_row + 3
        ws.cell(row=sig_row, column=2, value="ลายเซ็นต์รับเงิน").alignment = center
        ws.cell(row=sig_row, column=8, value="วันที่").alignment = center
        for c in range(1, 11):
            ws.cell(row=sig_row, column=c).border = thin

        col_widths = [14, 20, 14, 14, 12, 8, 12, 10, 12, 15]
        for i, width in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = width
    
    output = BytesIO()
    wb.save(output)
    return output.getvalue()


def main():
    if len(sys.argv) < 3:
        print(json.dumps({"error": "Usage: python excel-generator.py <command> <data_file>"}))
        sys.exit(1)

    command = sys.argv[1]
    data_file = sys.argv[2]

    try:
        # Read data from file
        with open(data_file, 'r', encoding='utf-8') as f:
            data = json.load(f)
        
        if command == "summary":
            excel_bytes = create_summary_excel(data["summaries"])
            print(base64.b64encode(excel_bytes).decode('utf-8'))

        elif command == "individual":
            excel_bytes = create_individual_excel(
                data["emp_name"],
                data["records"],
                data.get("period_text", "")
            )
            print(base64.b64encode(excel_bytes).decode('utf-8'))

        elif command == "all":
            excel_bytes = create_all_employees_excel(
                data["employees"],
                data.get("period_text", "")
            )
            print(base64.b64encode(excel_bytes).decode('utf-8'))

        else:
            print(json.dumps({"error": f"Unknown command: {command}"}))
            sys.exit(1)

    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)


if __name__ == "__main__":
    main()
