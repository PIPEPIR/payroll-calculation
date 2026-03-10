import re
import io
import math
import streamlit as st
import pandas as pd
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter

# regex จับแต่ละบรรทัดข้อมูลในไฟล์ PDF ที่มีรูปแบบ:
# StaffID  ชื่อพนักงาน  วันที่-เวลา  ประเภทการตอกบัตร
_ROW_RE = re.compile(
    r'^(\d+)\s+(.+?)\s+(\d{4}[/\-]\d{2}[/\-]\d{2}\s+\d{2}:\d{2}:\d{2})\s+(.+)$'
)

st.set_page_config(page_title="ระบบคิดเงินเดือนร้านอาหาร", page_icon="📝", layout="wide")

# ==========================================
# Global CSS
# ==========================================
st.markdown(
    """
    <style>
    /* ซ่อนคำแนะนำ "Press Enter to apply" */
    div[data-testid="InputInstructions"] > span:nth-child(1) {
        visibility: hidden;
        display: none;
    }
    div[data-testid="InputInstructions"] {
        visibility: hidden;
        position: absolute;
    }

    /* Wide layout */
    .block-container { max-width: 95%; }

    /* Progress Stepper */
    .step-bar {
        display: flex;
        align-items: center;
        margin-bottom: 1.5rem;
        font-size: 0.95rem;
    }
    .step-item {
        display: flex;
        align-items: center;
        gap: 6px;
    }
    .step-circle {
        width: 28px;
        height: 28px;
        border-radius: 50%;
        display: flex;
        align-items: center;
        justify-content: center;
        font-weight: bold;
        font-size: 0.85rem;
        flex-shrink: 0;
    }
    .step-circle.done   { background: #28a745; color: white; }
    .step-circle.active { background: #FF6B00; color: white; }
    .step-circle.todo   { background: #dee2e6; color: #6c757d; }
    .step-label.active  { font-weight: 600; color: #FF6B00; }
    .step-label.done    { color: #28a745; }
    .step-label.todo    { color: #6c757d; }
    .step-connector {
        flex: 0 0 40px;
        height: 2px;
        margin: 0 6px;
    }
    .step-connector.done { background: #28a745; }
    .step-connector.todo { background: #dee2e6; }
    </style>
    """,
    unsafe_allow_html=True,
)

# ==========================================
# Session state initialization
# ==========================================
for _key, _default in [
    ("step", 1),
    ("uploader_key", 0),
    ("raw_df", None),
    ("settings_df", None),
]:
    if _key not in st.session_state:
        st.session_state[_key] = _default


# ==========================================
# Helpers
# ==========================================
def render_stepper(current: int):
    steps = [
        (1, "อัปโหลด PDF"),
        (2, "ตั้งค่าค่าจ้าง"),
        (3, "ผลลัพธ์ & ส่งออก"),
    ]
    parts = []
    for i, (num, label) in enumerate(steps):
        if num < current:
            circle_cls, label_cls, icon = "done", "done", "✓"
        elif num == current:
            circle_cls, label_cls, icon = "active", "active", str(num)
        else:
            circle_cls, label_cls, icon = "todo", "todo", str(num)

        parts.append(
            f'<div class="step-item">'
            f'<div class="step-circle {circle_cls}">{icon}</div>'
            f'<span class="step-label {label_cls}">{label}</span>'
            f'</div>'
        )
        if i < len(steps) - 1:
            conn_cls = "done" if current > num else "todo"
            parts.append(f'<div class="step-connector {conn_cls}"></div>')

    st.markdown(f'<div class="step-bar">{"".join(parts)}</div>', unsafe_allow_html=True)


# ==========================================
# Sidebar — How to Use
# ==========================================
with st.sidebar:
    st.header("📖 วิธีใช้งาน")
    st.markdown("""
**ขั้นตอนการใช้งาน:**

1. **อัปโหลด PDF** — ไฟล์จากเครื่องตอกบัตรมาตรฐาน (รองรับหลายไฟล์)
2. **ตั้งค่าค่าจ้าง** — ระบุอัตราค่าจ้างพื้นฐานรายวันของแต่ละคน
3. **ตรวจสอบ & ส่งออก** — ดูผลและดาวน์โหลด Excel

---

**รูปแบบ PDF ที่รองรับ:**
- ออกจากเครื่องตอกบัตรรุ่นมาตรฐาน
- แต่ละแถวประกอบด้วย:
  รหัส / ชื่อ / วันที่-เวลา / ประเภทบันทึก

---

**ติดต่อสอบถาม:**
หากพบปัญหา กรุณาแจ้งผู้ดูแลระบบ
    """)

# ==========================================
# Page title
# ==========================================
st.title("📝 ระบบคิดเงินเดือน")


# ==========================================
# Core functions (data / business logic)
# ==========================================
@st.cache_data(show_spinner=False)
def read_file_to_df(file) -> pd.DataFrame:
    file_bytes = file.read()

    import fitz  # PyMuPDF
    doc = fitz.open(stream=file_bytes, filetype="pdf")

    lines = []
    for page in doc:
        text = page.get_text("text", sort=True)
        if text:
            lines.extend(text.splitlines())
    doc.close()

    if not lines:
        raise ValueError("ไม่พบข้อมูลในไฟล์ PDF")

    rows = []
    for line in lines:
        m = _ROW_RE.match(line.strip())
        if m:
            staff_id, full_name, timestamp, record_type = m.groups()
            rows.append({
                "Employee Staff ID": staff_id.strip(),
                "Full Name": full_name.strip(),
                "Timestamp": timestamp.strip(),
                "Record Type": record_type.strip(),
            })

    if not rows:
        raise ValueError(
            "ไม่พบข้อมูลแถวในไฟล์ — กรุณาตรวจสอบว่าไฟล์ถูก export "
            "จากเครื่องตอกบัตรในรูปแบบที่รองรับ"
        )

    return pd.DataFrame(rows)


def build_settings_df(raw_df: pd.DataFrame) -> pd.DataFrame:
    settings = []
    for name, _ in raw_df.groupby("Full Name"):
        settings.append({
            "ชื่อพนักงาน": name.strip(),
            "ค่าจ้างต่อวัน (บาท)": 300,
        })
    return pd.DataFrame(settings)


def calculate_employee_payroll(df: pd.DataFrame, base_daily_rate: int, staff_id: str = "", emp_name: str = ""):
    df = df[["Timestamp"]].copy()
    df["Timestamp"] = pd.to_datetime(df["Timestamp"], format="mixed")
    df = df.sort_values("Timestamp").reset_index(drop=True)
    df["Date"] = df["Timestamp"].dt.date

    daily_records = []
    total_days = 0
    total_hours = 0.0
    total_penalty = 0
    auto_checkout_days = []
    forgot_checkin_days = []

    for date, group in df.groupby("Date"):
        raw_punches = group["Timestamp"].tolist()

        # กรองตอกบัตรเบิ้ล (ถ้าห่างกันไม่เกิน 5 นาที ให้ข้าม)
        punches = []
        for p in raw_punches:
            if not punches:
                punches.append(p)
            elif (p - punches[-1]).total_seconds() > 300:
                punches.append(p)

        if not punches:
            continue

        first_punch = punches[0]

        # กำหนดกะอัตโนมัติ: สแกนเข้าก่อน 15:45 → กะ 14:00 | 15:45+ → กะ 16:00 ได้ 250฿
        if first_punch.hour > 15 or (first_punch.hour == 15 and first_punch.minute >= 45):
            shift_start_hour = 16
            daily_rate = 250
        else:
            shift_start_hour = 14
            daily_rate = base_daily_rate

        shift_start = first_punch.replace(hour=shift_start_hour, minute=0, second=0, microsecond=0)

        # สายเกิน 4 ชม. → ถือว่าลืมตอกบัตรเข้างาน
        if (first_punch - shift_start).total_seconds() / 60 > 240:
            forgot_checkin_days.append(str(date))
            daily_records.append({
                "รหัสพนักงาน": staff_id, "ชื่อ": emp_name, "วันที่": date,
                "เวลาเข้างาน": first_punch.strftime("%H:%M"), "คิดเป็นเงิน": 0,
                "ทิป": 0, "รวมเงิน": 0, "เบิกเงิน": "", "เหลือสุทธิ": 0,
                "สาย (นาที)": "-", "โดนหัก (บาท)": 0, "ชั่วโมงทำงานรวม": 0,
                "หมายเหตุ": "⚠️ ลืมตอกบัตรเข้างาน - ข้ามการคำนวณ",
            })
            continue

        # Auto-checkout: ตอกบัตรเป็นเลขคี่ → เติม 23:59:59
        if len(punches) % 2 != 0:
            punches.append(punches[-1].replace(hour=23, minute=59, second=59, microsecond=0))
            auto_checkout_days.append(str(date))

        late_mins = 0
        daily_penalty = 0
        if first_punch > shift_start:
            late_mins = math.floor((first_punch - shift_start).total_seconds() / 60)
            daily_penalty = (late_mins * 5) if late_mins <= 30 else (30 * 5 + (late_mins - 30) * 10)

        total_penalty += daily_penalty

        daily_hours = round(sum(
            (punches[i + 1] - punches[i]).total_seconds() / 3600
            for i in range(0, len(punches) - 1, 2)
        ), 2)
        total_hours += daily_hours
        total_days += 1

        net_pay_today = daily_rate - daily_penalty
        daily_records.append({
            "รหัสพนักงาน": staff_id, "ชื่อ": emp_name, "วันที่": date,
            "เวลาเข้างาน": first_punch.strftime("%H:%M"), "คิดเป็นเงิน": net_pay_today,
            "ทิป": 0, "รวมเงิน": net_pay_today, "เบิกเงิน": "", "เหลือสุทธิ": net_pay_today,
            "สาย (นาที)": late_mins, "โดนหัก (บาท)": daily_penalty,
            "ชั่วโมงทำงานรวม": daily_hours,
            "หมายเหตุ": "⚠️ เติม checkout 23:59:59 อัตโนมัติ" if str(date) in auto_checkout_days else "",
        })

    return daily_records, total_days, round(total_hours, 2), total_penalty, auto_checkout_days, forgot_checkin_days


@st.cache_data(show_spinner=False)
def to_excel(df: pd.DataFrame) -> bytes:
    output = io.BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="สรุปยอดจ่ายเงิน")
        ws = writer.sheets["สรุปยอดจ่ายเงิน"]

        header_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
        header_font = Font(bold=True, color="FF000000", size=12)
        center = Alignment(horizontal="center", vertical="center")
        thin = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        for col_num, _ in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin

        for col_num, _ in enumerate(df.columns, 1):
            col_letter = get_column_letter(col_num)
            max_len = 0
            for cell in ws[col_letter]:
                try:
                    max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
                cell.border = thin
                if col_num > 1:
                    cell.alignment = center
            ws.column_dimensions[col_letter].width = max_len + 5

    return output.getvalue()


def to_individual_excel(emp_name: str, daily_records: list, period_text: str = "") -> bytes:
    """สร้าง Excel รายบุคคลตามเทมเพลตร้าน (9 คอลัมน์ + header 3 แถว)"""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "เงินเดือน"

    header_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
    header_font = Font(bold=True, color="FF000000", size=11)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    ws.merge_cells("A1:G1")
    ws.cell(row=1, column=1, value=f"เงินเดือนพนักงานติดมันส์สาขาหาดใหญ่ {emp_name}").font = title_font
    ws.merge_cells("A2:G2")
    ws.cell(row=2, column=1, value="โทร: .......................................")
    ws.merge_cells("A3:C3")
    ws.cell(row=3, column=1, value=f"วันที่ {period_text}" if period_text else "วันที่ ........................")
    ws.cell(row=3, column=5, value="เงินสด")

    col_headers = ["รหัสพนักงาน", "ชื่อ", "วันที่", "เวลาเข้างาน", "คิดเป็นเงิน", "ทิป", "รวมเงิน", "เบิกเงิน", "เหลือสุทธิ"]
    for col_num, header in enumerate(col_headers, 1):
        cell = ws.cell(row=5, column=col_num, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = center
        cell.border = thin

    last_row = 5
    for row_idx, rec in enumerate(daily_records, 6):
        last_row = row_idx
        date_val = rec.get("วันที่", "")
        if hasattr(date_val, "strftime"):
            date_val = date_val.strftime("%d/%m/%Y")
        values = [
            rec.get("รหัสพนักงาน", ""), rec.get("ชื่อ", ""), date_val,
            rec.get("เวลาเข้างาน", ""), rec.get("คิดเป็นเงิน", 0),
            rec.get("ทิป", 0), f"=E{row_idx}+F{row_idx}",
            rec.get("เบิกเงิน", ""), f"=G{row_idx}-H{row_idx}",
        ]
        for col_num, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_num, value=val)
            cell.border = thin
            if col_num >= 3:
                cell.alignment = center

    summary_row = last_row + 1
    if last_row >= 6:
        sum_base = f"=SUM(E6:E{last_row})"
        sum_tip = f"=SUM(F6:F{last_row})"
        sum_total = f"=SUM(G6:G{last_row})"
        sum_advance = f"=SUM(H6:H{last_row})"
        sum_net = f"=SUM(I6:I{last_row})"
    else:
        sum_base = sum_tip = sum_total = sum_advance = sum_net = 0

    ws.cell(row=summary_row, column=2, value=f"รวม {period_text}".strip()).alignment = center
    for c_idx, val in zip([5, 6, 7, 8, 9], [sum_base, sum_tip, sum_total, sum_advance, sum_net]):
        ws.cell(row=summary_row, column=c_idx, value=val).border = thin
    for c in [1, 2, 3, 4]:
        ws.cell(row=summary_row, column=c).border = thin

    for r in [summary_row + 1, summary_row + 2]:
        for c in range(1, 10):
            ws.cell(row=r, column=c).border = thin

    sig_row = summary_row + 3
    ws.cell(row=sig_row, column=2, value="ลายเซ็นต์รับเงิน").alignment = center
    ws.cell(row=sig_row, column=7, value="วันที่").alignment = center
    for c in range(1, 10):
        ws.cell(row=sig_row, column=c).border = thin

    for i, w in enumerate([14, 20, 14, 14, 12, 8, 12, 10, 12], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


def to_all_employees_custom_excel(employees_data: list, period_text: str = "") -> bytes:
    """สร้าง Excel รวมทุกคน แต่ละคนแยกตาม 1 Sheet"""
    from openpyxl import Workbook

    wb = Workbook()
    header_fill = PatternFill(start_color="FFFFA500", end_color="FFFFA500", fill_type="solid")
    header_font = Font(bold=True, color="FF000000", size=11)
    title_font = Font(bold=True, size=13)
    center = Alignment(horizontal="center", vertical="center")
    thin = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for idx, emp_data in enumerate(employees_data):
        emp_name = emp_data["emp_name"]
        daily_records = emp_data["records"]
        sheet_title = str(emp_name).replace("/", "-").replace("\\", "-")[:31] or f"Employee_{idx+1}"

        ws = wb.active if idx == 0 else wb.create_sheet(title=sheet_title)
        if idx == 0:
            ws.title = sheet_title

        ws.merge_cells("A1:G1")
        ws.cell(row=1, column=1, value=f"เงินเดือนพนักงานติดมันส์สาขาหาดใหญ่ {emp_name}").font = title_font
        ws.merge_cells("A2:G2")
        ws.cell(row=2, column=1, value="โทร: .......................................")
        ws.merge_cells("A3:C3")
        ws.cell(row=3, column=1, value=f"วันที่ {period_text}" if period_text else "วันที่ ........................")
        ws.cell(row=3, column=5, value="เงินสด")

        col_headers = ["รหัสพนักงาน", "ชื่อ", "วันที่", "เวลาเข้างาน", "คิดเป็นเงิน", "ทิป", "รวมเงิน", "เบิกเงิน", "เหลือสุทธิ"]
        for col_num, header in enumerate(col_headers, 1):
            cell = ws.cell(row=5, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = center
            cell.border = thin

        last_row = 5
        for row_idx, rec in enumerate(daily_records, 6):
            last_row = row_idx
            date_val = rec.get("วันที่", "")
            if hasattr(date_val, "strftime"):
                date_val = date_val.strftime("%d/%m/%Y")
            values = [
                rec.get("รหัสพนักงาน", ""), rec.get("ชื่อ", ""), date_val,
                rec.get("เวลาเข้างาน", ""), rec.get("คิดเป็นเงิน", 0),
                rec.get("ทิป", 0), f"=E{row_idx}+F{row_idx}",
                rec.get("เบิกเงิน", ""), f"=G{row_idx}-H{row_idx}",
            ]
            for col_num, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_num, value=val)
                cell.border = thin
                if col_num >= 3:
                    cell.alignment = center

        summary_row = last_row + 1
        if last_row >= 6:
            sum_base = f"=SUM(E6:E{last_row})"
            sum_tip = f"=SUM(F6:F{last_row})"
            sum_total = f"=SUM(G6:G{last_row})"
            sum_advance = f"=SUM(H6:H{last_row})"
            sum_net = f"=SUM(I6:I{last_row})"
        else:
            sum_base = sum_tip = sum_total = sum_advance = sum_net = 0

        ws.cell(row=summary_row, column=2, value=f"รวม {period_text}".strip()).alignment = center
        for c_idx, val in zip([5, 6, 7, 8, 9], [sum_base, sum_tip, sum_total, sum_advance, sum_net]):
            ws.cell(row=summary_row, column=c_idx, value=val).border = thin
        for c in [1, 2, 3, 4]:
            ws.cell(row=summary_row, column=c).border = thin

        for r in [summary_row + 1, summary_row + 2]:
            for c in range(1, 10):
                ws.cell(row=r, column=c).border = thin

        sig_row = summary_row + 3
        ws.cell(row=sig_row, column=2, value="ลายเซ็นต์รับเงิน").alignment = center
        ws.cell(row=sig_row, column=7, value="วันที่").alignment = center
        for c in range(1, 10):
            ws.cell(row=sig_row, column=c).border = thin

        for i, w in enumerate([14, 20, 14, 14, 12, 8, 12, 10, 12], 1):
            ws.column_dimensions[get_column_letter(i)].width = w

    output = io.BytesIO()
    wb.save(output)
    return output.getvalue()


# ==========================================
# Step 1: Upload PDF
# ==========================================
if st.session_state.step == 1:
    render_stepper(1)
    st.header("อัปโหลดไฟล์ PDF ข้อมูลตอกบัตร")

    st.info(
        "**รูปแบบไฟล์ที่รองรับ:** PDF ที่ออกจากเครื่องตอกบัตรมาตรฐาน "
        "โดยแต่ละแถวในไฟล์ต้องมีข้อมูลในรูปแบบ: "
        "**รหัสพนักงาน / ชื่อ / วันที่และเวลา / ประเภทการบันทึก**\n\n"
        "สามารถอัปโหลดหลายไฟล์พร้อมกันได้ ระบบจะนำข้อมูลมารวมกันโดยอัตโนมัติ"
    )

    uploaded_files = st.file_uploader(
        "ลากไฟล์มาวางหรือคลิกเพื่อเลือก (รองรับหลายไฟล์)",
        type=["pdf"],
        accept_multiple_files=True,
        key=f"uploader_{st.session_state.uploader_key}",
    )

    if uploaded_files:
        with st.spinner("กำลังอ่านข้อมูลจากไฟล์..."):
            try:
                all_dfs = [read_file_to_df(f) for f in uploaded_files]
                combined = pd.concat(all_dfs, ignore_index=True)
                st.session_state.raw_df = combined
                st.session_state.settings_df = build_settings_df(combined)
                st.session_state.step = 2
                st.rerun()
            except Exception as e:
                err = str(e)
                if "ไม่พบข้อมูล" in err:
                    st.error(
                        f"{err}\n\n"
                        "กรุณาตรวจสอบว่าไฟล์ PDF ถูก export มาจากเครื่องตอกบัตร "
                        "ในรูปแบบที่ถูกต้อง ไม่ใช่ไฟล์ที่สแกนมาจากกระดาษ"
                    )
                else:
                    st.error(
                        "อ่านไฟล์ไม่สำเร็จ — กรุณาตรวจสอบว่า:\n"
                        "- ไฟล์เป็น PDF ที่ออกจากเครื่องตอกบัตร (ไม่ใช่ PDF สแกน)\n"
                        "- ไฟล์ไม่ได้ถูกล็อกหรือเข้ารหัสไว้\n\n"
                        f"รายละเอียดข้อผิดพลาด: {err}"
                    )

# ==========================================
# Step 2: Configure wages
# ==========================================
elif st.session_state.step == 2:
    render_stepper(2)
    st.header("ตั้งค่าค่าจ้างพนักงาน")

    with st.expander("ℹ️ ดูเงื่อนไขการคำนวณและปรับกะ", expanded=False):
        st.markdown("""
**กฎการกำหนดกะอัตโนมัติ (รายวัน):**
- เข้างานก่อน 15:45 → **กะ 14:00** ได้รับค่าจ้างตามอัตราพื้นฐาน
- เข้างานตั้งแต่ 15:45 เป็นต้นไป → **กะ 16:00** ได้รับ **250 บาท** (ตายตัว)

**กฎการหักเงินสาย:**
- สายไม่เกิน 30 นาที → หักนาทีละ **5 บาท**
- สายเกิน 30 นาที → ส่วนที่เกิน 30 นาทีแรก หักนาทีละ **10 บาท**

**กรณีพิเศษ:**
- ตอกบัตรไม่ครบคู่ → ระบบเติม checkout เป็น **23:59:59** อัตโนมัติ
- สแกนซ้ำห่างกันไม่เกิน 5 นาที → ถือว่าตอกเดียวกัน (กรองออก)
- เวลาเข้างานเกินกำหนดเกิน 4 ชั่วโมง → ถือว่าลืมตอกบัตรเข้างาน ข้ามการคำนวณ
        """)

    # --- Bulk wage setter ---
    st.markdown("**ตั้งค่าจ้างพื้นฐานพร้อมกันทุกคน**")
    col_bulk, col_apply = st.columns([3, 1])
    with col_bulk:
        bulk_wage = st.number_input(
            "ค่าจ้างพื้นฐาน (บาท)",
            min_value=0, step=50, value=300,
            label_visibility="collapsed",
        )
    with col_apply:
        if st.button("ใช้กับทุกคน", use_container_width=True):
            st.session_state.settings_df["ค่าจ้างต่อวัน (บาท)"] = bulk_wage
            st.rerun()

    st.caption("หรือแก้ไขค่าจ้างรายคนในตารางด้านล่าง")

    edited_df = st.data_editor(
        st.session_state.settings_df,
        use_container_width=True,
        hide_index=True,
        num_rows="dynamic",
        column_config={
            "ชื่อพนักงาน": st.column_config.TextColumn("ชื่อพนักงาน", disabled=True),
            "ค่าจ้างต่อวัน (บาท)": st.column_config.NumberColumn(
                "ค่าจ้างต่อวันพื้นฐาน (บาท)",
                min_value=0, step=50, required=True,
            ),
        },
        key="settings_editor",
    )

    col_calc, col_back = st.columns([3, 1])
    with col_calc:
        if st.button("✅ ยืนยันและคำนวณเงินเดือน", type="primary", use_container_width=True):
            st.session_state.settings_df = edited_df
            st.session_state.step = 3
            st.rerun()
    with col_back:
        if st.button("🔙 กลับไปอัปโหลดใหม่", use_container_width=True):
            st.session_state.step = 1
            st.session_state.raw_df = None
            st.session_state.settings_df = None
            st.session_state.uploader_key += 1
            st.rerun()

# ==========================================
# Step 3: Results & Export
# ==========================================
elif st.session_state.step == 3:
    render_stepper(3)
    st.header("ผลการคำนวณเงินเดือน")

    raw_df = st.session_state.raw_df
    settings_df = st.session_state.settings_df
    rate_lookup = dict(zip(settings_df["ชื่อพนักงาน"], settings_df["ค่าจ้างต่อวัน (บาท)"]))

    # --- Process ALL employees first ---
    all_employees_summary = []
    all_employees_detail = []     # { emp_name, staff_id, records, total_days, total_hours, total_penalty, base_pay, net_pay }
    all_auto_checkout_warnings = []  # { name, dates }
    all_forgot_checkin_warnings = [] # { name, dates }

    employee_groups = [(name.strip(), grp) for name, grp in raw_df.groupby("Full Name")]

    with st.spinner("กำลังคำนวณ..."):
        for emp_name, emp_df in employee_groups:
            if emp_name not in rate_lookup:
                continue
            emp_rate = int(rate_lookup[emp_name])
            emp_staff_id = str(emp_df["Employee Staff ID"].iloc[0]) if "Employee Staff ID" in emp_df.columns else ""

            daily_records, total_days, total_hours, total_penalty, auto_checkout_days, forgot_checkin_days = \
                calculate_employee_payroll(emp_df, emp_rate, staff_id=emp_staff_id, emp_name=emp_name)

            if auto_checkout_days:
                all_auto_checkout_warnings.append({"name": emp_name, "dates": auto_checkout_days})
            if forgot_checkin_days:
                all_forgot_checkin_warnings.append({"name": emp_name, "dates": forgot_checkin_days})

            base_pay = total_days * emp_rate
            net_pay = base_pay - total_penalty

            all_employees_summary.append({
                "ชื่อพนักงาน": emp_name,
                "วันที่ทำงาน (วัน)": total_days,
                "ชั่วโมงทำงาน (ชม.)": total_hours,
                "ค่าจ้างปกติ (บาท)": base_pay,
                "หักมาสาย (บาท)": total_penalty,
                "รับเงินสุทธิ (บาท)": net_pay,
            })

            if daily_records:
                all_employees_detail.append({
                    "emp_name": emp_name,
                    "staff_id": emp_staff_id,
                    "records": daily_records,
                    "total_days": total_days,
                    "total_hours": total_hours,
                    "total_penalty": total_penalty,
                    "base_pay": base_pay,
                    "net_pay": net_pay,
                })

    if not all_employees_summary:
        st.warning("ไม่พบข้อมูลพนักงาน กรุณาตรวจสอบการตั้งค่า")
    else:
        summary_df = pd.DataFrame(all_employees_summary)
        grand_total = summary_df["รับเงินสุทธิ (บาท)"].sum()
        total_penalty_all = summary_df["หักมาสาย (บาท)"].sum()
        headcount = len(summary_df)

        # ── Row 1: KPI metrics ──
        st.markdown("### สรุปภาพรวม")
        k1, k2, k3 = st.columns(3)
        k1.metric("ยอดเงินรวมที่ต้องโอนจ่าย", f"฿{grand_total:,.0f}")
        k2.metric("จำนวนพนักงาน", f"{headcount} คน")
        k3.metric("ยอดหักสายรวม", f"฿{total_penalty_all:,.0f}")

        st.write("")

        # ── Row 2: Aggregated warnings ──
        if all_forgot_checkin_warnings:
            total_missing_days = sum(len(w["dates"]) for w in all_forgot_checkin_warnings)
            with st.expander(
                f"📛 {len(all_forgot_checkin_warnings)} คน ลืมตอกบัตรเข้างาน (มีแต่ตอกออก) "
                f"(รวม {total_missing_days} วัน) — คลิกเพื่อดูรายละเอียด",
                expanded=True,
            ):
                for w in all_forgot_checkin_warnings:
                    st.markdown(f"- **{w['name']}**: {', '.join(w['dates'])}")
                st.error("ระบบข้ามการคำนวณรายได้ของวันข้างต้น: ไม่พบข้อมูลตอกบัตรเข้างาน (หรือเข้างานสายเกิน 4 ชั่วโมง) โปรดตรวจสอบเวลาเข้างานอีกครั้ง")

        if all_auto_checkout_warnings:
            total_issue_days = sum(len(w["dates"]) for w in all_auto_checkout_warnings)
            with st.expander(
                f"⚠️ {len(all_auto_checkout_warnings)} คน มีวันตอกบัตรไม่ครบคู่ "
                f"(รวม {total_issue_days} วัน) — คลิกเพื่อดูรายละเอียด",
                expanded=False,
            ):
                for w in all_auto_checkout_warnings:
                    st.markdown(f"- **{w['name']}**: {', '.join(w['dates'])}")
                st.caption("ระบบเติม checkout เป็น 23:59:59 อัตโนมัติ กรุณาตรวจสอบชั่วโมงทำงานในไฟล์ Excel")

        # ── Row 3: Summary table ──
        st.markdown("### ตารางสรุปพนักงานทั้งหมด")
        st.dataframe(summary_df, use_container_width=True, hide_index=True)

        st.write("")

        # ── Row 4: Per-employee detail (dropdown) ──
        if all_employees_detail:
            st.markdown("### ดูรายละเอียดรายวันของพนักงาน")
            emp_names_list = [e["emp_name"] for e in all_employees_detail]
            selected_emp = st.selectbox(
                "เลือกพนักงาน",
                options=emp_names_list,
                index=0,
                label_visibility="collapsed",
            )
            emp_detail = next(e for e in all_employees_detail if e["emp_name"] == selected_emp)

            dc1, dc2, dc3, dc4 = st.columns(4)
            dc1.metric("วันทำงาน", f"{emp_detail['total_days']} วัน")
            dc2.metric("ชั่วโมงรวม", f"{emp_detail['total_hours']:.1f} ชม.")
            dc3.metric("หักสาย", f"฿{emp_detail['total_penalty']:,.0f}")
            dc4.metric("รับสุทธิ", f"฿{emp_detail['net_pay']:,.0f}")

            df_daily = pd.DataFrame(emp_detail["records"])
            ui_cols = ["วันที่", "เวลาเข้างาน", "สาย (นาที)", "โดนหัก (บาท)", "ชั่วโมงทำงานรวม", "คิดเป็นเงิน", "หมายเหตุ"]
            display_cols = [c for c in ui_cols if c in df_daily.columns]
            st.dataframe(df_daily[display_cols], use_container_width=True, hide_index=True)

        st.write("")
        st.divider()

        # ── Row 5: Export Center ──
        st.markdown("### ส่งออกข้อมูล")
        ex1, ex2 = st.columns(2)
        with ex1:
            if all_employees_detail:
                st.download_button(
                    label="📥 ดาวน์โหลดรายบุคคล (แยก Sheet ทุกคน)",
                    data=to_all_employees_custom_excel(all_employees_detail),
                    file_name=f"payroll_all_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    type="primary",
                    use_container_width=True,
                )
        with ex2:
            st.download_button(
                label="📥 ดาวน์โหลดสรุปยอดรวม (ตารางเดียว)",
                data=to_excel(summary_df),
                file_name=f"payroll_summary_{pd.Timestamp.now().strftime('%Y%m%d_%H%M')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True,
            )

    st.write("")
    if st.button("🔄 เริ่มใหม่", type="primary"):
        st.session_state.step = 1
        st.session_state.raw_df = None
        st.session_state.settings_df = None
        st.session_state.uploader_key += 1
        st.rerun()
